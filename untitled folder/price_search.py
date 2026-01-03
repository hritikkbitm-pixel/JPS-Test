#!/usr/bin/env python3
"""
Automated Laptop Price Search Script
Searches Amazon and Flipkart for laptop prices and updates Excel file with final-price column.
"""

import openpyxl
from openpyxl.styles import PatternFill
import re
import time
import urllib.parse
import json
import subprocess
import sys

# File paths
INPUT_FILE = '/Users/Hritik/Desktop/JPS-Test/untitled folder/T3 NLC- 1st Jan to 31st Jan.xlsx'
OUTPUT_FILE = '/Users/Hritik/Desktop/JPS-Test/untitled folder/T3 NLC- 1st Jan to 31st Jan - WITH PRICES.xlsx'

# Yellow fill for cells where price was not found
YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# Rate limiting delay between searches (seconds)
SEARCH_DELAY = 3

def extract_price_from_text(text):
    """Extract price from text, looking for Indian Rupee amounts."""
    if not text:
        return None
    
    # Look for patterns like ₹45,999 or Rs. 45,999 or Rs 45999 or 45,999
    patterns = [
        r'₹\s*([\d,]+(?:\.\d{2})?)',
        r'Rs\.?\s*([\d,]+(?:\.\d{2})?)',
        r'INR\s*([\d,]+(?:\.\d{2})?)',
        r'(\d{2,3},\d{3}(?:\.\d{2})?)',  # Pattern like 45,999 or 1,45,999
    ]
    
    prices = []
    for pattern in patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        for match in matches:
            try:
                # Remove commas and convert to float
                price = float(match.replace(',', ''))
                # Filter out unrealistic prices (too low or too high)
                if 10000 <= price <= 500000:
                    prices.append(price)
            except ValueError:
                continue
    
    return min(prices) if prices else None

def search_amazon_price(search_query):
    """Search Amazon India for product price using browser automation."""
    try:
        encoded_query = urllib.parse.quote(search_query)
        url = f"https://www.amazon.in/s?k={encoded_query}"
        
        # Use AppleScript to get page content from browser
        script = f'''
        tell application "System Events"
            -- Open URL in default browser
            do shell script "open '{url}'"
            delay 5
        end tell
        '''
        
        # Use curl with proper headers and decompress gzip
        result = subprocess.run([
            'curl', '-s', '-L', '--compressed',
            '-H', 'User-Agent: Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            '-H', 'Accept-Language: en-IN,en;q=0.9',
            '-H', 'Accept-Encoding: gzip, deflate',
            f'https://www.amazon.in/s?k={encoded_query}'
        ], capture_output=True, text=True, timeout=30)
        
        if result.returncode == 0:
            # Look for price patterns in the HTML
            html = result.stdout
            
            # Amazon price patterns
            price_patterns = [
                r'class="a-price-whole"[^>]*>([^<]+)<',
                r'"priceAmount":(\d+(?:\.\d{2})?)',
                r'a-price[^>]*>.*?(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)',
                r'₹\s*(\d{1,3}(?:,\d{2,3})*(?:\.\d{2})?)',
            ]
            
            for pattern in price_patterns:
                matches = re.findall(pattern, html)
                for match in matches:
                    try:
                        price = float(match.replace(',', '').replace(' ', ''))
                        if 10000 <= price <= 500000:
                            return price
                    except (ValueError, AttributeError):
                        continue
        
        return None
    except Exception as e:
        print(f"  Amazon search error: {e}")
        return None

def search_flipkart_price(search_query):
    """Search Flipkart for product price."""
    try:
        encoded_query = urllib.parse.quote(search_query)
        
        result = subprocess.run([
            'curl', '-s', '-L',
            '-H', 'User-Agent: Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            '-H', 'Accept-Language: en-IN,en;q=0.9',
            f'https://www.flipkart.com/search?q={encoded_query}'
        ], capture_output=True, text=True, timeout=30)
        
        if result.returncode == 0:
            html = result.stdout
            
            # Flipkart price patterns
            price_patterns = [
                r'class="_30jeq3[^"]*"[^>]*>₹([^<]+)<',
                r'₹\s*(\d{1,3}(?:,\d{2,3})*(?:\.\d{2})?)',
                r'"selling_price":(\d+)',
                r'"price":(\d+)',
            ]
            
            for pattern in price_patterns:
                matches = re.findall(pattern, html)
                for match in matches:
                    try:
                        price = float(match.replace(',', '').replace(' ', ''))
                        if 10000 <= price <= 500000:
                            return price
                    except (ValueError, AttributeError):
                        continue
        
        return None
    except Exception as e:
        print(f"  Flipkart search error: {e}")
        return None

def build_search_query(row_data):
    """Build optimized search query from laptop data."""
    series = row_data.get('series', '') or ''
    mtm = row_data.get('mtm', '') or ''
    cpu = row_data.get('cpu', '') or ''
    ram_hdd = row_data.get('ram_hdd', '') or ''
    
    # Primary search: Lenovo + Series + Model Number (most precise)
    if mtm:
        return f"Lenovo {series} {mtm}".strip()
    
    # Fallback: Series + CPU + RAM
    query_parts = ['Lenovo', series]
    if cpu:
        query_parts.append(cpu)
    if ram_hdd:
        # Extract just RAM size
        ram_match = re.search(r'(\d+GB)', ram_hdd)
        if ram_match:
            query_parts.append(ram_match.group(1))
    
    return ' '.join(filter(None, query_parts))

def get_cell_value(ws, row, col):
    """Get cell value, handling formulas."""
    cell = ws.cell(row=row, column=col)
    if cell.value is not None and str(cell.value).startswith('='):
        # If it's a formula, try to get the cached value
        return cell.value
    return cell.value

def main():
    print("=" * 60)
    print("AUTOMATED LAPTOP PRICE SEARCH")
    print("=" * 60)
    print(f"\nInput file: {INPUT_FILE}")
    print(f"Output file: {OUTPUT_FILE}")
    
    # Load workbook with data_only=True to get calculated values
    print("\nLoading Excel file (may take a moment)...")
    
    # First, load with formulas to preserve structure
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb.active
    
    # Also load with data_only to get calculated values
    wb_data = openpyxl.load_workbook(INPUT_FILE, data_only=True)
    ws_data = wb_data.active
    
    # Find last row with data
    last_row = ws.max_row
    print(f"Found {last_row - 4} laptop entries (rows 5 to {last_row})")
    
    # Add header for the new column X (final-price)
    ws.cell(row=4, column=24, value="final-price")
    
    # Column mappings
    COL_B = 2   # Series
    COL_C = 3   # Form Factor
    COL_D = 4   # MTM (Model Number)
    COL_E = 5   # CPU
    COL_F = 6   # AI PC
    COL_G = 7   # RAM/HDD
    COL_H = 8   # Operating System
    COL_I = 9   # Office
    COL_J = 10  # Graphics
    COL_W = 23  # Effective T3 NLC taxpaid (cost price)
    COL_X = 24  # final-price (new column)
    
    results = {
        'total': 0,
        'price_found': 0,
        'price_not_found': 0,
        'errors': 0
    }
    
    print("\n" + "-" * 60)
    print("Starting price search...")
    print("-" * 60)
    
    for row in range(5, last_row + 1):
        results['total'] += 1
        
        # Get row data
        series = ws.cell(row=row, column=COL_B).value
        mtm = ws.cell(row=row, column=COL_D).value
        cpu = ws.cell(row=row, column=COL_E).value
        ram_hdd = ws.cell(row=row, column=COL_G).value
        
        # Get cost price from Column W (use data_only version for calculated values)
        cost_price = ws_data.cell(row=row, column=COL_W).value
        
        # Skip empty rows
        if not series and not mtm:
            continue
        
        # Try to convert cost_price to float
        try:
            if cost_price is not None:
                cost_price = float(cost_price)
            else:
                cost_price = 0
        except (ValueError, TypeError):
            cost_price = 0
        
        row_data = {
            'series': series,
            'mtm': mtm,
            'cpu': cpu,
            'ram_hdd': ram_hdd
        }
        
        search_query = build_search_query(row_data)
        
        print(f"\n[{results['total']}/{last_row - 4}] {series} - {mtm}")
        print(f"  Search: {search_query}")
        print(f"  Cost (W): ₹{cost_price:,.2f}" if cost_price else "  Cost (W): N/A")
        
        # Search Amazon and Flipkart
        print("  Searching Amazon...", end=" ", flush=True)
        amazon_price = search_amazon_price(search_query)
        print(f"₹{amazon_price:,.2f}" if amazon_price else "Not found")
        
        time.sleep(SEARCH_DELAY)
        
        print("  Searching Flipkart...", end=" ", flush=True)
        flipkart_price = search_flipkart_price(search_query)
        print(f"₹{flipkart_price:,.2f}" if flipkart_price else "Not found")
        
        # Determine final price
        prices = [p for p in [amazon_price, flipkart_price] if p is not None]
        
        final_price_cell = ws.cell(row=row, column=COL_X)
        
        if prices:
            lowest_price = min(prices)
            final_price = lowest_price - 100
            results['price_found'] += 1
            print(f"  ✓ Lowest: ₹{lowest_price:,.2f} → Final: ₹{final_price:,.2f}")
        else:
            # Price not found - use cost + 2% and highlight
            if cost_price > 0:
                final_price = cost_price * 1.02
                final_price_cell.fill = YELLOW_FILL
                results['price_not_found'] += 1
                print(f"  ⚠ Price not found. Using cost + 2%: ₹{final_price:,.2f}")
            else:
                final_price = 0
                final_price_cell.fill = YELLOW_FILL
                results['errors'] += 1
                print(f"  ✗ Error: No price or cost available")
        
        final_price_cell.value = round(final_price, 2) if final_price else 0
        
        time.sleep(SEARCH_DELAY)
    
    # Save the workbook
    print("\n" + "=" * 60)
    print("SAVING RESULTS")
    print("=" * 60)
    
    wb.save(OUTPUT_FILE)
    print(f"\n✓ Saved to: {OUTPUT_FILE}")
    
    # Print summary
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"Total laptops processed: {results['total']}")
    print(f"Prices found online:     {results['price_found']}")
    print(f"Prices not found:        {results['price_not_found']} (highlighted yellow)")
    print(f"Errors:                  {results['errors']}")
    print("\nDone!")

if __name__ == '__main__':
    main()
